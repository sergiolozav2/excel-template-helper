import os
import sys
import json
import io
from openpyxl import Workbook, load_workbook 
from openpyxl.utils.cell import get_column_letter, range_boundaries

from datetime import datetime

def is_valid_iso_date(date_string):
    try:
        datetime.strptime(date_string, "%Y-%m-%dT%H:%M:%S.%fZ")
        return True
    except:
        return False

def log_to_file(message):
    log_file = "debug_logs.txt"
    with open(log_file, "a") as f:
        f.write(message + "\n")

def open_template(filename, sheet):
    wb = load_workbook(filename)
    ws = wb[sheet]
    return wb, ws

try:
    request = json.loads(sys.stdin.read())

    if not request:
        sys.exit(0)

    if request.get("template_path") is None:
        raise ValueError("'template_path is required in the request.")
    if request.get("sheet") is None:
        raise ValueError("'sheet' is required in the request.")
    if request.get("data") is None:
        raise ValueError("'data' is required in the request.")

    if not isinstance(request["data"], list):
        raise ValueError("'data' must be a list of dictionaries.")

    (wb, ws) = open_template(request["template_path"], request["sheet"])  # Fixed attribute access

    start_row = request.get("start_row", 1)
    start_col = request.get("start_col", 1)

    for row_index, rows in enumerate(request["data"]):
        for col_index, value in enumerate(rows):

            if is_valid_iso_date(value):
                try:
                    value = datetime.strptime(value, "%Y-%m-%dT%H:%M:%S.%fZ")
                except Exception as e:
                    pass

            ws.cell(row=start_row + row_index, column=start_col + col_index, value=value)
    
    table_name = request.get("table_name", None)
    if(table_name):
        if(table_name not in ws.tables):
            raise ValueError(f"Table '{table_name}' does not exist.")
        table = ws.tables[table_name]

        table_height = len(request["data"])

        if table_height == 0:
            raise ValueError("No data provided to fill the table.")
            sys.exit(0)
        new_table_end = get_column_letter(range_boundaries(table.ref)[2]) + str(start_row + table_height - 1)
        new_ref = f"{table.ref.split(':')[0]}:{new_table_end}"
        table.ref = new_ref

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    sys.stdout.buffer.write(buffer.read())

except Exception as e:
    print(f"Python Error: {e}", file=sys.stderr)
    sys.exit(1)


